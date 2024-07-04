
import streamlit as st
import pandas as pd
import math
from collections import defaultdict
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def get_first_digit(number):
    return int(str(number)[0])
    
def get_first_two_digit(number):
    return int(str(number)[:2])

def get_first_three_digit(number):
    return int(str(number)[:3])

def benfords_law(data):
    counts = defaultdict(int)
    total = 0
    for value in data:
        first_digit = get_first_digit(value)
        counts[first_digit] += 1
        total += 1
    
    benfords = [(math.log10(1 + 1 / digit) * total) for digit in range(1, 10)]
    observed = [counts[digit] for digit in range(1, 10)]
    return benfords, observed

def benfords1_law(data, i):
    counts = defaultdict(int)
    total = 0
    for value in data:
        first_two_digits = get_first_two_digit(value)
        counts[first_two_digits] += 1
        total += 1
    
    benfords = [math.log10(1 + 1 / digit) * total for digit in range(10 * i + 1, 10 * i + 10)]
    observed = [counts[digit] for digit in range(10 * i + 1, 10 * i + 10)]
    return benfords, observed

# Set page title and favicon
st.set_page_config(page_title="Benford's Law Analysis", page_icon=":bar_chart:")

# Set page title
st.title("Benford's Law Analysis")

# Upload Excel file
uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

if uploaded_file is not None:
    # Load Excel file
    df = pd.read_excel(uploaded_file)
    df2 = df.copy()
    # Select column for analysis
    column_to_analyze = st.selectbox("Select column for analysis", options=df.columns)
    data = df[column_to_analyze].dropna().astype(int).values

    x = len(data)
    benfords, observed = benfords_law(data)
    deviations = [((observed[i] - benfords[i]) / x) * 100 for i in range(9)]

    # Display Benford's law deviations
    st.subheader("Benford's Law Deviation:")
    for digit, deviation in zip(range(1, 10), deviations):
        st.write(f"Digit {digit}: Deviation {deviation:.2f}%")
    st.set_option('deprecation.showPyplotGlobalUse', False)

    # Separate the positive and negative deviations
    lstneg, lstpos = defaultdict(int), defaultdict(int)
    for digit, deviation in zip(range(1, 10), deviations):
        if deviation < 0:
            lstneg[deviation] = digit
        elif deviation > 0:
            lstpos[deviation] = digit

    # Sort the dictionaries
    lstpos = sorted(lstpos.items(), reverse=True)
    lstneg = sorted(lstneg.items())

    # Plot the subplots
    maxdev = defaultdict(int)
    for key, value in lstpos:
        benfords, observed = benfords1_law(data, value)
        deviations = [((observed[i] - benfords[i]) / x) * 100 for i in range(9)]
        listx = [i for i in range(10 * value + 1, 10 * value + 10)]
        plt.plot(listx, deviations)
        plt.title("Benford Distribution of Values Exceeding Requirements")
        for i in range(len(deviations)):
            if deviations[i] > 0:
                maxdev[deviations[i]] = 10 * value + i

    color = 'FFFF00'  
    plt.legend()
    st.pyplot()

    maxdev = sorted(maxdev.items())
    st.subheader("Benford's Law Analysis Based on 2nd Digit:")
    for key, value in maxdev:
        st.write(f"Deviation {key:.2f}%: Digit {value}")

    # Load the workbook and the first worksheet
    wb = load_workbook(uploaded_file)
    ws = wb.active
    column_index = df2.columns.get_loc(column_to_analyze) + 1

    # Apply color fill to cells with specific starting digits
    for key, val in maxdev:
        for r_idx, value in enumerate(df2[column_to_analyze], start=2):  # Start from row 2 to match Excel's 1-based indexing
            cell = ws.cell(row=r_idx, column=column_index)
            if str(value).startswith(str(val)) and r_idx > 1:  # Check if it's a data row
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Save the modified file
    modified_file_path = 'modified_file.xlsx'
    wb.save(modified_file_path)

    # Provide download link for the modified file
    with open(modified_file_path, "rb") as file:
        st.download_button(label="Download Modified Excel File", data=file, file_name=modified_file_path)

